import datetime
import os
import io
import hashlib

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, FileResponse
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from docx import Document as DocxDocument
from docx.shared import Pt, RGBColor

from .models import (
    User, Role, Project, ProjectStatus, Task, TaskStatus,
    Document, DocumentType, Comment, ProjectMessage,
    TaskHistory, Lead, Notification
)

def notify(user, message, link=None):
    if user:
        Notification.objects.create(user=user, message=message, link=link)

def index(request):
    return redirect('projects')

def login_view(request):
    if request.method == "POST":
        u = request.POST.get('username')
        p = request.POST.get('password')
        # In Django usually you would authenticate, but if their current passwords
        # are just SHA256 hashed without salt (like get_password_hash in fastapi),
        # standard authenticate won't work out of the box unless we setup custom Auth backend.
        # For this prototype we will bypass or authenticate directly if we manually migrate.
        user = User.objects.filter(username=u).first()
        if user and (user.check_password(p) or (hasattr(user, 'password_hash') and user.password_hash == hashlib.sha256(p.encode()).hexdigest())):
            if hasattr(user, 'password_hash') and user.password_hash == hashlib.sha256(p.encode()).hexdigest() and not user.has_usable_password():
                 user.set_password(p)
                 user.save()
            
            auth_login(request, user)
            rn = user.role.name if user.role else ""
            if rn == "employee":
                return redirect('workspace')
            return redirect('projects')
        
        return render(request, 'login.html', {"error": "Неверный логин или пароль"})
    return render(request, 'login.html')

def logout_view(request):
    auth_logout(request)
    return redirect('login')

@login_required(login_url='login')
def profile(request):
    if request.method == "POST":
        request.user.full_name = request.POST.get('full_name', '')
        request.user.email = request.POST.get('email', '')
        request.user.phone = request.POST.get('phone', '')
        request.user.position = request.POST.get('position', '')
        request.user.save()
        return redirect('/profile/?saved=1')
    return render(request, 'profile.html', {'user': request.user})

@login_required(login_url='login')
def list_projects(request):
    rn = request.user.role.name if request.user.role else ""
    if rn in ("head", "pm"):
        projects = Project.objects.all()
    elif rn == "customer":
        projects = Project.objects.filter(customer=request.user)
    else:
        # employee - redirect to workspace or show limited
        projects = Project.objects.filter(tasks__assignee=request.user).distinct()
    return render(request, 'projects.html', {'user': request.user, 'projects': projects})

@login_required(login_url='login')
def employee_workspace(request):
    """Unified workspace for an employee."""
    user = request.user
    # 1. Statistics
    tasks = Task.objects.filter(assignee=user)
    total_tasks = tasks.count()
    done_tasks = tasks.filter(status__name='Выполнено').count()
    active_tasks = tasks.exclude(status__name='Выполнено').count()
    
    import datetime
    now = datetime.date.today()
    overdue_tasks = tasks.filter(deadline__lt=now).exclude(status__name='Выполнено').count()
    
    # 2. My Projects
    projects = Project.objects.filter(tasks__assignee=user).distinct()
    
    # 3. Recent notifications
    notifications = Notification.objects.filter(user=user, is_read=False).order_by('-created_at')[:5]
    
    return render(request, 'employee_workspace.html', {
        'user': user,
        'total_tasks': total_tasks,
        'done_tasks': done_tasks,
        'active_tasks': active_tasks,
        'overdue_tasks': overdue_tasks,
        'projects': projects,
        'notifications': notifications,
        'now': now
    })

@login_required(login_url='login')
def new_project(request):
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('projects')
    
    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description')
        customer_id = request.POST.get('customer_id')
        pm_id = request.POST.get('pm_id') or request.user.id
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date') or None
        planned_budget = request.POST.get('planned_budget', 0.0)
        
        status, _ = ProjectStatus.objects.get_or_create(name="Новый")
        
        proj = Project.objects.create(
            title=title, description=description,
            customer_id=customer_id, pm_id=pm_id,
            start_date=start_date, end_date=end_date,
            planned_budget=float(planned_budget), status=status
        )
        if proj.customer:
            notify(proj.customer, f"Создан проект «{title}».", f"/projects/{proj.id}")
        return redirect('projects')
        
    customers = User.objects.filter(role__name="customer")
    pms = User.objects.filter(role__name__in=["pm", "head"])
    return render(request, 'project_form.html', {'user': request.user, 'customers': customers, 'pms': pms})

@login_required(login_url='login')
def gantt_chart(request):
    """View for the Project Gantt Chart."""
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('projects')
        
    projects = Project.objects.all().order_by('start_date')
    
    # Simple progress calculation: (completed tasks / total tasks) * 100
    gantt_data = []
    for p in projects:
        total_tasks = p.tasks.count()
        completed_tasks = p.tasks.filter(status__name='Выполнено').count()
        progress = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        gantt_data.append({
            'id': str(p.id),
            'name': p.title,
            'start': p.start_date.isoformat(),
            'end': (p.end_date or (p.start_date + datetime.timedelta(days=30))).isoformat(),
            'progress': int(progress),
            'dependencies': ''
        })
        
    import json
    return render(request, 'gantt.html', {
        'user': request.user,
        'projects_json': json.dumps(gantt_data)
    })

@login_required(login_url='login')
def project_detail(request, project_id):
    proj = get_object_or_404(Project, id=project_id)
    rn = request.user.role.name if request.user.role else ""
    if rn == "customer" and proj.customer_id != request.user.id:
        return redirect('projects')
        
    Notification.objects.filter(user=request.user, link=f"/projects/{project_id}", is_read=False).update(is_read=True)
    doc_types = DocumentType.objects.all()
    return render(request, 'project_detail.html', {'user': request.user, 'project': proj, 'doc_types': doc_types})

@login_required(login_url='login')
def project_message(request, project_id):
    if request.method == "POST":
        proj = get_object_or_404(Project, id=project_id)
        content = request.POST.get('content')
        is_revision = request.POST.get('is_revision') == 'true'
        
        if content:
            ProjectMessage.objects.create(project=proj, sender=request.user, content=content, is_revision=is_revision)
            rn = request.user.role.name if request.user.role else ""
            if rn == "customer" and proj.pm:
                notify(proj.pm, f"Новое сообщение от заказчика в проекте «{proj.title}»", f"/projects/{project_id}")
            elif proj.customer:
                notify(proj.customer, f"Новое сообщение по проекту «{proj.title}»", f"/projects/{project_id}")
            
    return redirect(f"/projects/{project_id}/")

@login_required(login_url='login')
def upload_document(request, project_id):
    if request.method == "POST":
        proj = get_object_or_404(Project, id=project_id)
        doc_title = request.POST.get('doc_title')
        doc_type_id = request.POST.get('doc_type_id')
        file = request.FILES.get('file')
        
        if file:
            if not doc_title:
                doc_title = file.name
            Document.objects.create(
                project=proj, title=doc_title, document_type_id=doc_type_id,
                file_path=file, author=request.user
            )
    return redirect(f"/projects/{project_id}/")

@login_required(login_url='login')
def download_doc(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    if doc.file_path:
        return FileResponse(doc.file_path.open('rb'), as_attachment=True, filename=os.path.basename(doc.file_path.name))
    return HttpResponse(status=404)

@login_required(login_url='login')
def list_tasks(request):
    rn = request.user.role.name if request.user.role else ""
    if rn in ("head", "pm"):
        tasks = Task.objects.all()
    elif rn == "employee":
        tasks = Task.objects.filter(assignee=request.user)
    else:
        return redirect('projects')
    import datetime
    now = datetime.date.today()
    return render(request, 'tasks.html', {'user': request.user, 'tasks': tasks, 'now': now})

@login_required(login_url='login')
def kanban_board(request):
    """View for the Kanban board."""
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm", "employee"):
        return redirect('projects')
        
    if rn == "employee":
        tasks = Task.objects.filter(assignee=request.user)
    else:
        tasks = Task.objects.all()
        
    statuses = TaskStatus.objects.all()
    status_order = ["К выполнению", "В процессе", "Выполнено"]
    ordered_statuses = []
    for s_name in status_order:
        s_obj = statuses.filter(name=s_name).first()
        if s_obj:
            ordered_statuses.append(s_obj)
            
    import datetime
    now = datetime.date.today()
    return render(request, 'kanban.html', {
        'user': request.user,
        'statuses': ordered_statuses,
        'tasks': tasks,
        'now': now
    })

@login_required(login_url='login')
def update_task_status_api(request, task_id):
    """AJAX API to update task status."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Post only'}, status=400)
        
    task = get_object_or_404(Task, id=task_id)
    status_id = request.POST.get('status_id')
    
    if not status_id:
        return JsonResponse({'error': 'No status_id'}, status=400)
        
    new_status = get_object_or_404(TaskStatus, id=status_id)
    
    # Save old status for history
    old_status = task.status
    task.status = new_status
    task.save()
    
    # Optional: Log to TaskHistory
    TaskHistory.objects.create(
        task=task,
        old_status=old_status,
        new_status=new_status,
        changed_by=request.user
    )
    
    return JsonResponse({'success': True})

@login_required(login_url='login')
def new_task(request):
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('tasks')
        
    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        project_id = request.POST.get('project_id')
        assignee_id = request.POST.get('assignee_id')
        deadline = request.POST.get('deadline')
        cost = request.POST.get('cost', 0.0)
        
        status, _ = TaskStatus.objects.get_or_create(name="К выполнению")
        task = Task.objects.create(
            title=title, description=description, project_id=project_id,
            assignee_id=assignee_id, deadline=deadline,
            cost=float(cost), status=status
        )
        if task.assignee:
            notify(task.assignee, f"Вам назначена задача: «{title}»", f"/projects/{project_id}")
        return redirect(request.META.get('HTTP_REFERER', 'tasks'))
        
    comp_status = ProjectStatus.objects.filter(name="Завершен").first()
    projects = Project.objects.exclude(status=comp_status) if comp_status else Project.objects.all()
    employees = User.objects.filter(role__name="employee")
    preselect_project = request.GET.get('project_id')
    return render(request, 'task_form.html', {
        'user': request.user, 'projects': projects,
        'employees': employees, 'preselect_project': preselect_project
    })

@login_required(login_url='login')
def complete_task(request, task_id):
    if request.method == "POST":
        task = get_object_or_404(Task, id=task_id)
        rn = request.user.role.name if request.user.role else ""
        if rn in ("head", "pm") or task.assignee == request.user:
            sd, _ = TaskStatus.objects.get_or_create(name="Выполнено")
            old_status = task.status
            task.status = sd
            task.save()
            TaskHistory.objects.create(task=task, old_status=old_status, new_status=sd, changed_by=request.user)
            if task.project and task.project.pm:
                notify(task.project.pm, f"Задача «{task.title}» выполнена исполнителем {request.user.full_name}", f"/projects/{task.project.id}")
    return redirect(request.META.get('HTTP_REFERER', 'tasks'))

@login_required(login_url='login')
def list_users(request):
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('projects')
    
    q = request.GET.get('q', '')
    sort = request.GET.get('sort', 'full_name')
    
    users = User.objects.all()
    if q:
        from django.db.models import Q
        users = users.filter(Q(username__icontains=q) | Q(full_name__icontains=q) | Q(position__icontains=q))
    
    # Sorting
    if sort == 'role':
        users = users.order_by('role__name', 'full_name')
    else:
        # Default sort by name
        users = users.order_by('full_name')

    employees = User.objects.filter(role__name="employee")
    active_status = TaskStatus.objects.filter(name__in=["К выполнению", "В процессе"])
    workload = {}
    for emp in employees:
        act = Task.objects.filter(assignee=emp, status__in=active_status).count()
        tot = Task.objects.filter(assignee=emp).count()
        workload[emp.id] = {"active": act, "total": tot}
        
    return render(request, 'users.html', {
        'user': request.user, 
        'users': users, 
        'workload': workload,
        'query': q,
        'current_sort': sort
    })

@login_required(login_url='login')
def list_leads(request):
    rn = request.user.role.name if request.user.role else ""
    if rn in ("head", "pm"):
        leads = Lead.objects.all()
    elif rn == "customer":
        leads = Lead.objects.filter(customer=request.user)
    else:
        return redirect('projects')
    doc_types = DocumentType.objects.all()
    return render(request, 'leads.html', {'user': request.user, 'leads': leads, 'doc_types': doc_types})

@login_required(login_url='login')
def new_lead(request):
    if request.method == "POST" and getattr(request.user.role, 'name', '') == "customer":
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        budget = request.POST.get('budget', 0.0)
        desired_deadline = request.POST.get('desired_deadline') or None
        tz_file = request.FILES.get('tz_file')
        
        lead = Lead.objects.create(
            title=title, description=description, budget=float(budget),
            desired_deadline=desired_deadline, customer=request.user, status="Новая"
        )
        if tz_file:
            lead.tz_file_path = tz_file
            lead.save()
            
        managers = User.objects.filter(role__name__in=["head", "pm"])
        for m in managers:
            notify(m, f"Новая заявка от {request.user.full_name}: «{title}»", "/leads")
    return redirect('leads')

@login_required(login_url='login')
def approve_lead(request, lead_id):
    if request.method == "POST" and getattr(request.user.role, 'name', '') in ("head", "pm"):
        lead = get_object_or_404(Lead, id=lead_id)
        if lead.status == "Новая":
            lead.status = "Одобрена"
            lead.save()
            sn, _ = ProjectStatus.objects.get_or_create(name="Новый")
            proj = Project.objects.create(
                title=lead.title, description=lead.description, customer=lead.customer,
                pm=request.user, start_date=datetime.date.today(), end_date=lead.desired_deadline,
                planned_budget=lead.budget, status=sn
            )
            notify(lead.customer, f"Ваша заявка «{lead.title}» одобрена — создан проект!", f"/projects/{proj.id}")
    return redirect('projects')

@login_required(login_url='login')
def reject_lead(request, lead_id):
    if request.method == "POST" and getattr(request.user.role, 'name', '') in ("head", "pm"):
        lead = get_object_or_404(Lead, id=lead_id)
        if lead.status == "Новая":
            lead.status = "Отклонена"
            lead.save()
            notify(lead.customer, f"Заявка «{lead.title}» отклонена. Обратитесь к менеджеру.", "/leads")
    return redirect('leads')

@login_required(login_url='login')
def notifications_page(request):
    notifs = Notification.objects.filter(user=request.user).order_by('-created_at')[:50]
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return render(request, 'notifications.html', {'user': request.user, 'notifs': notifs})

@login_required(login_url='login')
def notif_count(request):
    cnt = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'count': cnt})

@login_required(login_url='login')
def export_excel(request, project_id):
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('projects')
    proj = get_object_or_404(Project, id=project_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета проекта"
    
    # Simple boilerplate
    ws["A1"] = f"СМЕТА ПРОЕКТА: {proj.title}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    fname = f"smeta_{proj.id}_{proj.title[:20].replace(' ','_')}.xlsx"
    return HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename="{fname}"'})

@login_required(login_url='login')
def export_word(request, project_id):
    rn = request.user.role.name if request.user.role else ""
    if rn not in ("head", "pm"):
        return redirect('projects')
    proj = get_object_or_404(Project, id=project_id)
    
    doc = DocxDocument()
    doc.add_heading(f"Техническое задание\n«{proj.title}»", 0)
    
    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    
    fname = f"tz_{proj.id}_{proj.title[:20].replace(' ','_')}.docx"
    return HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document', headers={'Content-Disposition': f'attachment; filename="{fname}"'})

def sync_employees_1c(request):
    users = User.objects.filter(role__name="employee")
    return JsonResponse({"status": "success", "data": [{"id": u.id, "name": u.full_name} for u in users]})
